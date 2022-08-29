# program for reading excel file named all_routine containing routine of all sections to make 
# other individual routines and then writing them to a new excel file
import os
from openpyxl import load_workbook, Workbook

all_routine = load_workbook('routine1.xlsx')
# getting active worksheet (first worksheet)
active_sheet = all_routine.active
all_groups = []
# creating a list of all groups
for i in range(1,15):
	all_groups.append(f"L6CG{i}")

# making folder for storing individual routine
# exception handling
try:
	os.mkdir("individual_routines")
	print("New directory called individual_routines created.")
except:
	print("Directory already exist continuing further.")

for group in all_groups:
	# creating a new workbook for each group
	wb = Workbook()
	ws = wb.active
	# title of worksheet
	ws.title = group + "_routine_3rdSem"

	routine_heading = []
	for cell in active_sheet[2]:

		# getting value of each cell of second row ( 'Day', 'Time'..)
		routine_heading.append(cell.value)
	
	# appending routine heading to the first row of the sheet
	ws.append(routine_heading)
	# optimizing for current routine
	gc = group + "+"
	if group[4:] in ["4", "8", "11", "14"]:
		gc = "L5CG+" + group[4:] + ")"
	elif group[4:] in ["2", "3", "6", "10", "13"]:
		gc = "L5CG+" + group[4:] + "+"
	# first two rows of the original routine contains unneccessary data
	for row in active_sheet.iter_rows(4, active_sheet.max_row):

		if (group == row[3].value) or ((gc[4:]) in row[3].value[4:]):
			# row that will contain all the information about the class
			new_row = []
			for i in range(active_sheet.max_column):

				# appending cell value to new row
				new_row.append(row[i].value)

			# appending each day routine to the new excel file
			ws.append(new_row)
			# saving file inside individual_routines folder
			wb.save(f"individual_routines/{group}_routine_3rdSem.xlsx")
	wb.close()
all_routine.close()
print("You can find all routines in individual_routines folder.")